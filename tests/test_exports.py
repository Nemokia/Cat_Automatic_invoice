"""
Tests for PDF generation, Excel export, and report endpoints.

PDF: /api/pdf/<invoice_id>/ → application/pdf
Excel: /api/export/excel/?type=invoices|items → xlsx
Reports: /api/reports/sales/, /api/reports/customers/, /api/reports/products/
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework import status
from openpyxl import load_workbook

from tests.factories import (
    create_user, create_user2, create_customer, create_product,
    create_bank, create_bank_account, create_invoice, create_price_history,
)


class TestPDFGeneration(TestCase):
    """PDF generation for valid invoices."""

    def setUp(self):
        self.client = APIClient()
        self.user = create_user()
        self.client.force_authenticate(user=self.user)
        self.customer = create_customer(self.user)
        self.product = create_product(self.user)
        self.bank = create_bank()
        self.bank_account = create_bank_account(self.user, bank=self.bank)
        self.invoice = create_invoice(
            self.user,
            customer=self.customer,
            items_data=[
                {'product_name': self.product.name, 'quantity': 2, 'unit_price': 50000, 'tax_rate': 10, 'unit': self.product.unit},
                {'product_name': 'سرویس فیلتر', 'quantity': 1, 'unit_price': 120000, 'tax_rate': 0, 'unit': 'عدد'},
            ],
        )
        self.url = f'/api/pdf/{self.invoice.pk}/'

    def test_valid_invoice_returns_pdf(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_pdf_content_type(self):
        response = self.client.get(self.url)
        self.assertEqual(response['Content-Type'], 'application/pdf',
                         "PDF response must have application/pdf content type")

    def test_pdf_non_empty_content(self):
        response = self.client.get(self.url)
        self.assertTrue(len(response.content) > 0, "PDF content must not be empty")

    def test_pdf_starts_with_percent_pdf(self):
        response = self.client.get(self.url)
        self.assertTrue(response.content[:4] == b'%PDF',
                        "PDF must start with %PDF header")


class TestPDFNotFound(TestCase):
    """PDF for nonexistent invoice → 404."""

    def setUp(self):
        self.client = APIClient()
        self.user = create_user()
        self.client.force_authenticate(user=self.user)

    def test_nonexistent_invoice_returns_404(self):
        response = self.client.get('/api/pdf/99999/')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)


class TestPDFIsolation(TestCase):
    """User A's invoice → User B gets 404."""

    def setUp(self):
        self.client = APIClient()
        self.user_a = create_user(username='user_a', email='a@test.com')
        self.user_b = create_user(username='user_b', email='b@test.com')
        self.customer_a = create_customer(self.user_a)
        self.invoice_a = create_invoice(
            self.user_a,
            customer=self.customer_a,
            items_data=[{'product_name': 'Item A', 'quantity': 1, 'unit_price': 10000}],
        )

    def test_user_b_cannot_access_user_a_invoice(self):
        self.client.force_authenticate(user=self.user_b)
        response = self.client.get(f'/api/pdf/{self.invoice_a.pk}/')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND,
                         "User B must not access User A's invoice PDF")

    def test_user_a_can_access_own_invoice(self):
        self.client.force_authenticate(user=self.user_a)
        response = self.client.get(f'/api/pdf/{self.invoice_a.pk}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class TestExcelExport(TestCase):
    """Excel export with invoices type."""

    def setUp(self):
        self.client = APIClient()
        self.user = create_user()
        self.client.login(username=self.user.username, password="TestPass123!")
        self.customer = create_customer(self.user)
        self.product = create_product(self.user)
        self.invoice = create_invoice(
            self.user,
            customer=self.customer,
            items_data=[
                {'product_name': self.product.name, 'quantity': 3, 'unit_price': 25000, 'tax_rate': 9, 'unit': self.product.unit},
            ],
        )
        self.url = '/api/export/excel/'

    def test_invoices_type_returns_xlsx(self):
        response = self.client.get(self.url, {'type': 'invoices'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            "Excel response must have proper xlsx content type",
        )

    def test_valid_xlsx_content(self):
        response = self.client.get(self.url, {'type': 'invoices'})
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertIsNotNone(ws, "Workbook should have an active sheet")

    def test_has_expected_headers(self):
        response = self.client.get(self.url, {'type': 'invoices'})
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected_headers = ['شماره فاکتور', 'مشتری', 'تلفن', 'تاریخ',
                            'جمع', 'مالیات اقلام', 'تخفیف', 'نهایی', 'وضعیت']
        for expected in expected_headers:
            self.assertIn(expected, headers, f"Missing header: {expected}")

    def test_data_row_present(self):
        response = self.client.get(self.url, {'type': 'invoices'})
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        # Row 1 is headers, row 2+ should be data
        self.assertGreaterEqual(ws.max_row, 2, "Should have at least one data row")


class TestExcelExportItems(TestCase):
    """Excel export with items type."""

    def setUp(self):
        self.client = APIClient()
        self.user = create_user()
        self.client.login(username=self.user.username, password="TestPass123!")
        self.customer = create_customer(self.user)
        self.product = create_product(self.user)
        self.invoice = create_invoice(
            self.user,
            customer=self.customer,
            items_data=[
                {'product_name': self.product.name, 'quantity': 5, 'unit_price': 30000, 'tax_rate': 0, 'unit': self.product.unit},
            ],
        )
        self.url = '/api/export/excel/'

    def test_items_type_returns_xlsx(self):
        response = self.client.get(self.url, {'type': 'items'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('application/vnd.openxmlformats', response['Content-Type'])

    def test_items_type_has_item_columns(self):
        response = self.client.get(self.url, {'type': 'items'})
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected_headers = ['شماره فاکتور', 'نام کالا', 'تعداد', 'قیمت واحد', 'مالیات', 'مبلغ کل']
        for expected in expected_headers:
            self.assertIn(expected, headers, f"Missing items header: {expected}")


class TestExcelExportIsolation(TestCase):
    """User A exports → User B's data not included."""

    def setUp(self):
        self.client = APIClient()
        self.user_a = create_user(username='user_a', email='a@test.com')
        self.user_b = create_user(username='user_b', email='b@test.com')
        self.customer_a = create_customer(self.user_a)
        self.customer_b = create_customer(self.user_b, first_name='حسین', last_name='رضايی', phone='09351112233')
        self.product_a = create_product(self.user_a, name='لنت عقب')
        self.product_b = create_product(self.user_b, name='روغن موتور')
        self.invoice_a = create_invoice(
            self.user_a,
            customer=self.customer_a,
            items_data=[{'product_name': self.product_a.name, 'quantity': 2, 'unit_price': 45000}],
        )
        self.invoice_b = create_invoice(
            self.user_b,
            customer=self.customer_b,
            items_data=[{'product_name': self.product_b.name, 'quantity': 1, 'unit_price': 80000}],
        )

    def test_user_a_export_excludes_user_b_data(self):
        self.client.login(username=self.user_a.username, password="TestPass123!")
        response = self.client.get('/api/export/excel/', {'type': 'invoices'})
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        # Check all data rows — should not contain user_b's customer name
        all_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_data.extend([str(v) for v in row if v is not None])
        combined = ' '.join(all_data)
        self.assertNotIn('حسین', combined,
                         "User A's export must not include User B's customer data")

    def test_user_b_export_excludes_user_a_data(self):
        self.client.login(username=self.user_b.username, password="TestPass123!")
        response = self.client.get('/api/export/excel/', {'type': 'invoices'})
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        all_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_data.extend([str(v) for v in row if v is not None])
        combined = ' '.join(all_data)
        self.assertNotIn('علی', combined,
                         "User B's export must not include User A's customer data")


class TestExcelFilteredExport(TestCase):
    """Excel export with date_from/date_to filters."""

    def setUp(self):
        self.client = APIClient()
        self.user = create_user()
        self.client.login(username=self.user.username, password="TestPass123!")
        self.customer = create_customer(self.user)
        self.product = create_product(self.user)

        # Invoice from today
        self.invoice_today = create_invoice(
            self.user,
            customer=self.customer,
            items_data=[{'product_name': 'Today Item', 'quantity': 1, 'unit_price': 50000}],
            invoice_date=date.today(),
        )
        # Invoice from 30 days ago
        self.invoice_old = create_invoice(
            self.user,
            customer=self.customer,
            items_data=[{'product_name': 'Old Item', 'quantity': 1, 'unit_price': 60000}],
            invoice_date=date.today() - timedelta(days=30),
        )
        # Invoice 60 days ago
        self.invoice_ancient = create_invoice(
            self.user,
            customer=self.customer,
            items_data=[{'product_name': 'Ancient Item', 'quantity': 1, 'unit_price': 70000}],
            invoice_date=date.today() - timedelta(days=60),
        )

    def test_filter_by_date_from(self):
        response = self.client.get('/api/export/excel/', {
            'type': 'invoices',
            'date_from': (date.today() - timedelta(days=31)).isoformat(),
        })
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        # Should include today + 30 days ago, but not 60 days ago
        all_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_data.extend([str(v) for v in row if v is not None])
        combined = ' '.join(all_data)
        # Check by invoice number (invoices-type export doesn't include product names)
        self.assertIn(self.invoice_today.invoice_number, combined)
        self.assertNotIn(self.invoice_ancient.invoice_number, combined,
                         "Ancient invoice should be filtered out by date_from")

    def test_filter_by_date_to(self):
        response = self.client.get('/api/export/excel/', {
            'type': 'invoices',
            'date_to': (date.today() - timedelta(days=1)).isoformat(),
        })
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        all_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_data.extend([str(v) for v in row if v is not None])
        combined = ' '.join(all_data)
        self.assertNotIn(self.invoice_today.invoice_number, combined,
                         "Today's invoice should be excluded by date_to")

    def test_filter_by_date_range(self):
        response = self.client.get('/api/export/excel/', {
            'type': 'invoices',
            'date_from': (date.today() - timedelta(days=31)).isoformat(),
            'date_to': date.today().isoformat(),
        })
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        all_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_data.extend([str(v) for v in row if v is not None])
        combined = ' '.join(all_data)
        self.assertIn(self.invoice_today.invoice_number, combined)
        self.assertNotIn(self.invoice_ancient.invoice_number, combined)


class TestReportsSales(TestCase):
    """Sales report returns summary with total_sales and invoice_count."""

    def setUp(self):
        self.client = APIClient()
        self.user = create_user()
        self.client.force_authenticate(user=self.user)
        self.customer = create_customer(self.user)
        self.product = create_product(self.user)
        self.invoice = create_invoice(
            self.user,
            customer=self.customer,
            items_data=[
                {'product_name': self.product.name, 'quantity': 2, 'unit_price': 100000, 'tax_rate': 0},
            ],
        )
        self.url = '/api/reports/sales/'

    def test_sales_report_returns_summary(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.data
        self.assertIn('summary', data, "Sales report must contain 'summary' key")
        summary = data['summary']
        self.assertIn('total_sales', summary, "Summary must contain 'total_sales'")
        self.assertIn('invoice_count', summary, "Summary must contain 'invoice_count'")

    def test_sales_report_invoice_count(self):
        response = self.client.get(self.url)
        summary = response.data['summary']
        self.assertEqual(summary['invoice_count'], 1,
                         "Invoice count should match created invoices")

    def test_sales_report_total_sales(self):
        response = self.client.get(self.url)
        summary = response.data['summary']
        self.assertGreater(summary['total_sales'], 0,
                           "Total sales should be > 0 for invoices with items")


class TestReportsCustomers(TestCase):
    """Customer report returns customer list with total_spent."""

    def setUp(self):
        self.client = APIClient()
        self.user = create_user()
        self.client.force_authenticate(user=self.user)
        self.customer = create_customer(self.user)
        self.product = create_product(self.user)
        self.invoice = create_invoice(
            self.user,
            customer=self.customer,
            items_data=[{'product_name': self.product.name, 'quantity': 3, 'unit_price': 50000}],
        )
        self.url = '/api/reports/customers/'

    def test_customer_report_returns_list(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list, "Customer report must return a list")

    def test_customer_report_has_total_spent(self):
        response = self.client.get(self.url)
        self.assertGreater(len(response.data), 0)
        customer_data = response.data[0]
        self.assertIn('total_spent', customer_data, "Customer must have total_spent")
        self.assertIn('name', customer_data, "Customer must have name")
        self.assertIn('num_invoices', customer_data, "Customer must have num_invoices")


class TestReportsProducts(TestCase):
    """Product report returns product list."""

    def setUp(self):
        self.client = APIClient()
        self.user = create_user()
        self.client.force_authenticate(user=self.user)
        self.customer = create_customer(self.user)
        self.product = create_product(self.user)
        # Create invoices with items linked to product via FK
        from invoices.models import Invoice, InvoiceItem, InvoiceNumberSequence
        for _ in range(2):
            inv = Invoice.objects.create(
                user=self.user,
                invoice_number=InvoiceNumberSequence.get_next_number(self.user),
                customer=self.customer,
                customer_name=self.customer.full_name,
                customer_phone=self.customer.phone,
                invoice_date=date(2026, 1, 1),
            )
            InvoiceItem.objects.create(
                invoice=inv, product=self.product,
                product_name=self.product.name, quantity=2,
                unit_price=75000, tax_rate=0, unit='عدد',
            )
            inv.calculate_totals()
            inv.save()
        self.url = '/api/reports/products/'

    def test_product_report_returns_list(self):
        """Product report returns a list of products with analytics."""
        response = self.client.get(self.url, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list)
        self.assertGreaterEqual(len(response.data), 1)

    def test_product_report_has_expected_fields(self):
        """Product report includes expected fields."""
        response = self.client.get(self.url, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        item = response.data[0]
        self.assertIn('id', item)
        self.assertIn('name', item)
        self.assertIn('latest_price', item)
        self.assertIn('times_sold', item)
        self.assertIn('total_revenue', item)

    def test_product_report_times_sold(self):
        """Product report reflects times_sold from invoice items."""
        response = self.client.get(self.url, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        item = response.data[0]
        self.assertEqual(item['times_sold'], 2)
