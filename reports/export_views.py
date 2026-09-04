import io
import zipfile
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from invoices.models import Invoice, InvoiceItem
from customers.models import Customer
from products.models import Product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Column definitions per section
# ---------------------------------------------------------------------------
# Each entry: (field_name, header_label)
CUSTOMER_COLUMNS = [
    ('first_name', 'نام'),
    ('last_name', 'نام خانوادگی'),
    ('phone', 'تلفن'),
    ('address', 'آدرس'),
    ('national_id', 'شناسه ملی'),
    ('invoice_count', 'تعداد فاکتور'),
    ('total_purchases', 'مجموع خرید'),
]

PRODUCT_COLUMNS = [
    ('name', 'نام'),
    ('unit', 'واحد'),
    ('frequency', 'دوره'),
    ('latest_price', 'قیمت'),
    ('total_sold', 'تعداد فروش'),
    ('total_revenue', 'مجموع درآمد'),
]

INVOICE_COLUMNS = [
    ('invoice_number', 'شماره فاکتور'),
    ('customer_name', 'مشتری'),
    ('customer_phone', 'تلفن'),
    ('invoice_date', 'تاریخ'),
    ('due_date', 'سررسید'),
    ('subtotal', 'جمع'),
    ('item_taxes_total', 'مالیات اقلام'),
    ('discount_amount', 'تخفیف'),
    ('final_amount', 'نهایی'),
    ('is_paid', 'وضعیت'),
]

ITEM_COLUMNS = [
    ('invoice_number', 'شماره فاکتور'),
    ('product_name', 'نام کالا'),
    ('quantity', 'تعداد'),
    ('unit_price', 'قیمت واحد'),
    ('tax_amount', 'مالیات'),
    ('total_price', 'مبلغ کل'),
]

SUMMARY_COLUMNS = [
    ('month', 'ماه'),
    ('count', 'تعداد فاکتور'),
    ('revenue', 'مجموع درآمد'),
]

SECTION_COLUMNS = {
    'customers': CUSTOMER_COLUMNS,
    'products': PRODUCT_COLUMNS,
    'invoices': INVOICE_COLUMNS,
    'items': ITEM_COLUMNS,
    'summary': SUMMARY_COLUMNS,
}

# Persian header labels for backwards compat
SECTION_HEADERS = {
    'customers': [c[1] for c in CUSTOMER_COLUMNS],
    'products': [c[1] for c in PRODUCT_COLUMNS],
    'invoices': [c[1] for c in INVOICE_COLUMNS],
    'items': [c[1] for c in ITEM_COLUMNS],
    'summary': [c[1] for c in SUMMARY_COLUMNS],
}


def _get_selected_columns(section_name, request):
    """Return list of (field, header) tuples for the section.

    Query-param format: ``{section}_columns=field1,field2``
    When no columns param is supplied, return *all* columns for backward compat.
    """
    param_key = f'{section_name}_columns'
    raw = request.GET.get(param_key, '').strip()
    all_cols = SECTION_COLUMNS.get(section_name, [])

    if not raw:
        return all_cols  # backward compatible: export everything

    selected_keys = [k.strip() for k in raw.split(',') if k.strip()]
    return [(f, h) for f, h in all_cols if f in selected_keys]


# ---------------------------------------------------------------------------
# Customer value helpers
# ---------------------------------------------------------------------------
def _customer_value(c, field):
    if field == 'first_name':
        return c.first_name
    if field == 'last_name':
        return c.last_name
    if field == 'phone':
        return c.phone
    if field == 'address':
        return c.address
    if field == 'national_id':
        return c.national_id
    if field == 'invoice_count':
        return c.invoice_count
    if field == 'total_purchases':
        return float(c.total_purchases)
    return ''


def _product_value(p, field):
    if field == 'name':
        return p.name
    if field == 'unit':
        return p.unit
    if field == 'frequency':
        return p.get_frequency_display()
    if field == 'latest_price':
        return float(p.latest_price)
    if field == 'total_sold':
        return p.total_sold
    if field == 'total_revenue':
        return float(p.total_revenue)
    return ''


def _invoice_value(inv, field):
    if field == 'invoice_number':
        return inv.invoice_number
    if field == 'customer_name':
        return inv.customer_name
    if field == 'customer_phone':
        return inv.customer_phone
    if field == 'invoice_date':
        return str(inv.invoice_date)
    if field == 'due_date':
        return str(inv.due_date) if inv.due_date else ''
    if field == 'subtotal':
        return float(inv.subtotal)
    if field == 'item_taxes_total':
        return float(inv.item_taxes_total)
    if field == 'discount_amount':
        return float(inv.discount_amount)
    if field == 'final_amount':
        return float(inv.final_amount)
    if field == 'is_paid':
        return 'پرداخت شده' if inv.is_paid else 'پرداخت نشده'
    return ''


def _item_value(inv, item, field):
    if field == 'invoice_number':
        return inv.invoice_number
    if field == 'product_name':
        return item.product_name
    if field == 'quantity':
        return float(item.quantity)
    if field == 'unit_price':
        return float(item.unit_price)
    if field == 'tax_amount':
        return float(item.tax_amount)
    if field == 'total_price':
        return float(item.total_price)
    return ''


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
@login_required
def export_excel(request):
    """Export selected sections to Excel — multi-sheet workbook.
    Supports ?sections= (new) and ?type= (legacy) formats.
    Per-section columns via ?{section}_columns=field1,field2
    """
    user = request.user
    sections = request.GET.getlist('sections')
    if not sections:
        legacy_type = request.GET.get('type', 'invoices')
        sections = [legacy_type]
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    invoices = Invoice.objects.filter(user=user)
    if date_from:
        invoices = invoices.filter(invoice_date__gte=date_from)
    if date_to:
        invoices = invoices.filter(invoice_date__lte=date_to)

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0B1849', end_color='0B1849', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

    def write_row(ws, row, data):
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border

    first_sheet = True

    # --- Customers ---
    if 'customers' in sections:
        cols = _get_selected_columns('customers', request)
        if cols:
            ws = wb.active if first_sheet else wb.create_sheet()
            first_sheet = False
            ws.title = 'مشتریان'
            write_header(ws, [h for _, h in cols])
            for row, c in enumerate(Customer.objects.filter(user=user), 2):
                write_row(ws, row, [_customer_value(c, f) for f, _ in cols])
            for col in range(1, len(cols) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

    # --- Products ---
    if 'products' in sections:
        cols = _get_selected_columns('products', request)
        if cols:
            ws = wb.active if first_sheet else wb.create_sheet()
            first_sheet = False
            ws.title = 'محصولات'
            write_header(ws, [h for _, h in cols])
            for row, p in enumerate(Product.objects.filter(user=user), 2):
                write_row(ws, row, [_product_value(p, f) for f, _ in cols])
            for col in range(1, len(cols) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

    # --- Invoices ---
    if 'invoices' in sections:
        cols = _get_selected_columns('invoices', request)
        if cols:
            ws = wb.active if first_sheet else wb.create_sheet()
            first_sheet = False
            ws.title = 'فاکتورها'
            write_header(ws, [h for _, h in cols])
            for row, inv in enumerate(invoices, 2):
                write_row(ws, row, [_invoice_value(inv, f) for f, _ in cols])
            for col in range(1, len(cols) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

    # --- Invoice Items (details) ---
    if 'items' in sections:
        cols = _get_selected_columns('items', request)
        if cols:
            ws = wb.active if first_sheet else wb.create_sheet()
            first_sheet = False
            ws.title = 'اقلام فاکتور'
            write_header(ws, [h for _, h in cols])
            row = 2
            for inv in invoices:
                for item in inv.items.all():
                    write_row(ws, row, [_item_value(inv, item, f) for f, _ in cols])
                    row += 1
            for col in range(1, len(cols) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

    # --- Sales Summary ---
    if 'summary' in sections:
        cols = _get_selected_columns('summary', request)
        if cols:
            ws = wb.active if first_sheet else wb.create_sheet()
            first_sheet = False
            ws.title = 'خلاصه فروش'
            write_header(ws, [h for _, h in cols])
            from django.db.models.functions import TruncMonth
            from django.db.models import Sum, Count
            monthly = (
                invoices
                .annotate(month=TruncMonth('invoice_date'))
                .values('month')
                .annotate(revenue=Sum('final_amount'), count=Count('id'))
                .order_by('-month')
            )
            for row, m in enumerate(monthly, 2):
                data = []
                for f, _ in cols:
                    if f == 'month':
                        data.append(str(m['month'].strftime('%Y-%m')) if m['month'] else '')
                    elif f == 'count':
                        data.append(m['count'])
                    elif f == 'revenue':
                        data.append(float(m['revenue']))
                    else:
                        data.append('')
                write_row(ws, row, data)
            for col in range(1, len(cols) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

    # If no sheets were created (empty workbook), create a blank sheet
    if first_sheet:
        wb.active.title = 'Sheet1'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cat_invoice_export.xlsx"'
    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# Customer invoices PDF ZIP export
# ---------------------------------------------------------------------------
@login_required
def export_customer_invoices_pdf(request, customer_id):
    """Generate a ZIP file containing PDFs for all invoices of a given customer.

    URL: /api/export/pdf/<customer_id>/
    Files are named like INV-2026-000001.pdf
    """
    user = request.user

    # Verify the customer belongs to this user
    try:
        customer = Customer.objects.get(id=customer_id, user=user)
    except Customer.DoesNotExist:
        return HttpResponse('مشتری یافت نشد', status=404)

    # Get all invoices for this customer, ordered by invoice_date
    invoices = Invoice.objects.filter(user=user, customer=customer).order_by(
        '-invoice_date', '-created_at'
    )

    if not invoices.exists():
        return HttpResponse('فاکتوری برای این مشتری یافت نشد', status=404)

    # Generate PDFs in memory and pack into a ZIP
    zip_buffer = io.BytesIO()

    # Import the PDF generation helper
    from reports.pdf_views import generate_pdf_content

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for inv in invoices:
            pdf_bytes = generate_pdf_content(inv)
            filename = f'{inv.invoice_number}.pdf'
            zf.writestr(filename, pdf_bytes)

    zip_buffer.seek(0)

    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    safe_name = f'{customer.first_name}_{customer.last_name}'.replace(' ', '_')
    response['Content-Disposition'] = (
        f'attachment; filename="invoices_{safe_name}.zip"'
    )
    return response
